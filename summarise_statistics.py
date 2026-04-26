"""
summarise_statistics.py
=======================
Reads an AgNW analysis summary CSV and produces a statistics table with
mean, standard deviation, median, and percentiles for length, diameter,
and aspect ratio.

Samples are grouped by SAMPLE IDENTITY ONLY — all magnifications of the
same sample are pooled together into one row.

Images where the detected magnification is below --min_mag (default 20x) are
excluded. This removes mis-parsed date digits such as 13x and 19x that appear
in filenames like "sample_13_1_100x.png".

Works with both:
  - Per-run summary  (all_images_summary.csv from nanowire_analysis.py)
  - Batch summary    (all_samples_summary.csv from run_batch_analysis.py)

Usage
-----
    python summarise_statistics.py all_samples_summary.csv
    python summarise_statistics.py all_samples_summary.csv --output stats.xlsx
    python summarise_statistics.py all_samples_summary.csv --min_mag 50
"""

import sys
import csv
import math
import re
import argparse
from pathlib import Path
from collections import defaultdict

import numpy as np

# ---- Sample / magnification helpers -----------------------------------------

SAMPLE_NAMES = {"a","b","c","d","e","f","g","alpha","beta","gamma"}


def _extract_sample(image_name: str) -> str:
    """Extract sample letter/name from an image filename stem."""
    stem = re.sub(r"\.(tif|tiff|png|jpg|jpeg)$", "", image_name, flags=re.IGNORECASE)
    stem2 = re.sub(r"(\d+)([xX])(?=\D|$)", r"\1x ", stem)
    stem2 = re.sub(r"([A-Ww])(\d)", r"\1 \2", stem2)
    stem2 = re.sub(r"(\d)([A-Wa-wY-Zy-z])", r"\1 \2", stem2)
    tokens = [t.strip().lower() for t in re.split(r"[\s_\-]+", stem2) if t.strip()]
    for tok in tokens:
        if tok in SAMPLE_NAMES:
            return tok.capitalize()
    return image_name   # fallback


def _extract_mag_num(image_name: str) -> int:
    """
    Extract magnification as an integer, taking the LARGEST numeric token
    followed by x. This avoids picking up date digits (e.g. 13 in
    sample_13_1_100x -> returns 100, not 13).
    Returns 0 if nothing found.
    """
    stem = re.sub(r"\.(tif|tiff|png|jpg|jpeg)$", "", image_name, flags=re.IGNORECASE)
    stem2 = re.sub(r"(\d+)([xX])(?=\D|$)", r"\1x ", stem)
    tokens = [t.strip() for t in re.split(r"[\s_\-]+", stem2) if t.strip()]
    mags = []
    for tok in tokens:
        if re.fullmatch(r"\d{2,5}x?", tok, re.IGNORECASE):
            mags.append(int(re.sub(r"x$", "", tok, flags=re.IGNORECASE)))
    return max(mags) if mags else 0


# ---- Statistics helpers ------------------------------------------------------

def _floats(rows: list, col: str) -> "np.ndarray":
    vals = []
    for r in rows:
        raw = r.get(col, "")
        if raw in ("", "nan", "None", None):
            continue
        try:
            v = float(raw)
            if not math.isnan(v):
                vals.append(v)
        except (ValueError, TypeError):
            pass
    return np.array(vals, dtype=float)


def _stats(arr: "np.ndarray") -> dict:
    if len(arr) == 0:
        return {k: "" for k in ["n","mean","sd","median","p10","p90","min","max"]}
    return {
        "n":      int(len(arr)),
        "mean":   round(float(np.mean(arr)),           3),
        "sd":     round(float(np.std(arr, ddof=1)),    3),
        "median": round(float(np.median(arr)),         3),
        "p10":    round(float(np.percentile(arr, 10)), 3),
        "p90":    round(float(np.percentile(arr, 90)), 3),
        "min":    round(float(arr.min()),              3),
        "max":    round(float(arr.max()),              3),
    }


# ---- Main table builder ------------------------------------------------------

def build_stats_table(input_csv: Path, min_mag: int = 20):
    with open(input_csv, newline="", encoding="utf-8") as f:
        all_rows = list(csv.DictReader(f))
    if not all_rows:
        raise ValueError("CSV is empty.")

    fields   = list(all_rows[0].keys())
    is_batch = all(c in fields for c in ["synthesis_date","sample","magnification"])

    # Remove truncated wires
    rows    = [r for r in all_rows if r.get("truncated","False") not in ("True", True)]
    n_trunc = len(all_rows) - len(rows)

    # Remove rows from images with implausible magnification
    def _mag_num(r):
        if is_batch and r.get("magnification"):
            raw = re.sub(r"x$","", r["magnification"], flags=re.IGNORECASE)
            try: return int(raw)
            except ValueError: pass
        return _extract_mag_num(r.get("image",""))

    rows_ok  = [r for r in rows if _mag_num(r) >= min_mag]
    n_bad    = len(rows) - len(rows_ok)
    rows     = rows_ok

    print(f"  Rows read             : {len(all_rows)}")
    print(f"  Truncated removed     : {n_trunc}")
    print(f"  Bad mag (<{min_mag}x) removed : {n_bad}")
    print(f"  Rows used             : {len(rows)}")

    # Helpers to extract identity fields
    def _sample(r):
        return r["sample"] if is_batch and r.get("sample") else _extract_sample(r.get("image",""))
    def _modality(r):
        return r.get("modality", r.get("mode","?")).upper()
    def _date(r):
        return r.get("synthesis_date","").replace("_","-") if is_batch else ""

    # Group by (date, sample, modality) — magnification intentionally excluded
    groups: dict = defaultdict(list)
    for r in rows:
        groups[(_date(r), _sample(r), _modality(r))].append(r)

    # Build output
    stat_suf = ["n","mean","sd","median","p10","p90","min","max"]
    id_cols  = (["synthesis_date","sample","modality"] if is_batch
                else ["sample","modality"])
    out_fn   = (id_cols
                + [f"length_um_{s}"    for s in stat_suf]
                + [f"diameter_nm_{s}"  for s in stat_suf]
                + [f"aspect_ratio_{s}" for s in stat_suf]
                + ["orientation_mean_deg","orientation_R",
                   "tip_to_tip_pct","n_wires_total","magnifications_included"])
    out_rows = []

    for (date, sample, modality), grp in sorted(groups.items()):
        row = {}
        if is_batch: row["synthesis_date"] = date
        row["sample"]   = sample
        row["modality"] = modality

        # Which magnifications were pooled
        mags = sorted({_mag_num(r) for r in grp if _mag_num(r) > 0})
        row["magnifications_included"] = ", ".join(f"{m}x" for m in mags)

        # Length
        s = _stats(_floats(grp, "length_um"))
        for sf in stat_suf: row[f"length_um_{sf}"] = s[sf]

        # Diameter + aspect ratio (SEM only)
        if "SEM" in modality:
            for col, prefix in [("diameter_nm","diameter_nm"),("aspect_ratio","aspect_ratio")]:
                s = _stats(_floats(grp, col))
                for sf in stat_suf: row[f"{prefix}_{sf}"] = s[sf]
        else:
            for prefix in ["diameter_nm","aspect_ratio"]:
                for sf in stat_suf: row[f"{prefix}_{sf}"] = ""

        # Orientation
        angles = _floats(grp, "wire_angle_deg")
        if len(angles) >= 5:
            phi = 2.0 * np.radians(angles)
            R   = float(np.sqrt(np.mean(np.cos(phi))**2 + np.mean(np.sin(phi))**2))
            mu  = float(np.degrees(np.arctan2(np.mean(np.sin(phi)),
                                               np.mean(np.cos(phi)))) / 2 % 180)
            row["orientation_mean_deg"] = round(mu, 1)
            row["orientation_R"]        = round(R,  4)
        else:
            row["orientation_mean_deg"] = ""
            row["orientation_R"]        = ""

        # Tip-to-tip
        n_ep = sum(1 for r in grp if str(r.get("ep_to_ep","")).lower() in ("true","1"))
        row["tip_to_tip_pct"]  = round(100.0 * n_ep / len(grp), 1)
        row["n_wires_total"]   = len(grp)

        out_rows.append(row)

    return out_fn, out_rows


# ---- CSV output --------------------------------------------------------------

def save_csv(fieldnames, rows, path: Path):
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        w.writeheader(); w.writerows(rows)
    print(f"  Saved: {path}")


# ---- Excel output ------------------------------------------------------------

def save_xlsx(fieldnames, rows, path: Path):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  openpyxl not installed — saving as CSV instead.")
        print("  Install with:  pip install openpyxl")
        save_csv(fieldnames, rows, path.with_suffix(".csv"))
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sample Statistics"

    hdr_fill   = PatternFill("solid", fgColor="2D6A8C")
    sub_fill   = PatternFill("solid", fgColor="3A85A8")
    id_fill    = PatternFill("solid", fgColor="D6E4F0")
    alt_fill   = PatternFill("solid", fgColor="EEF6FB")
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    thin       = Side(style="thin", color="AAAAAA")
    border     = Border(left=thin, right=thin, top=thin, bottom=thin)
    centre     = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_al    = Alignment(horizontal="left",   vertical="center")
    hdr_font   = Font(bold=True, color="FFFFFF", size=10)
    sub_font   = Font(bold=True, color="FFFFFF", size=9)
    body_font  = Font(size=10)

    id_cols = [c for c in fieldnames if not any(
        c.startswith(p) for p in
        ["length_","diameter_","aspect_","orientation_","tip_","n_wires","mag"])]

    sections = [
        ("Sample",
         id_cols),
        ("Length (um)",
         [c for c in fieldnames if c.startswith("length_um_")]),
        ("Diameter (nm)",
         [c for c in fieldnames if c.startswith("diameter_nm_")]),
        ("Aspect ratio",
         [c for c in fieldnames if c.startswith("aspect_ratio_")]),
        ("Orientation & summary",
         [c for c in fieldnames
          if c.startswith(("orientation_","tip_","n_wires","mag"))]),
    ]
    sections = [(lbl, cols) for lbl, cols in sections if cols]

    col_order = []
    for _, cols in sections: col_order.extend(cols)

    NICE = {
        "synthesis_date":"Synthesis date","sample":"Sample","modality":"Modality",
        "image":"Image","mode":"Mode","magnifications_included":"Magnifications pooled",
        "n":"N","mean":"Mean","sd":"Std dev","median":"Median",
        "p10":"P10","p90":"P90","min":"Min","max":"Max",
        "orientation_mean_deg":"Orient. mean (deg)","orientation_R":"Orient. R",
        "tip_to_tip_pct":"Tip-to-tip (%)","n_wires_total":"N wires",
    }

    # Row 1 - section headers
    col = 1
    for lbl, cols in sections:
        cell = ws.cell(row=1, column=col, value=lbl)
        cell.font=hdr_font; cell.fill=hdr_fill
        cell.alignment=centre; cell.border=border
        if len(cols) > 1:
            ws.merge_cells(start_row=1, start_column=col,
                           end_row=1,   end_column=col+len(cols)-1)
        col += len(cols)

    # Row 2 - column headers
    for ci, cn in enumerate(col_order, 1):
        label = NICE.get(cn, NICE.get(cn.split("_")[-1], cn))
        cell  = ws.cell(row=2, column=ci, value=label)
        cell.font=sub_font; cell.fill=sub_fill
        cell.alignment=centre; cell.border=border

    # Data rows
    for ri, rd in enumerate(rows):
        er   = ri + 3
        fill = alt_fill if ri % 2 == 0 else white_fill
        for ci, cn in enumerate(col_order, 1):
            val = rd.get(cn, "")
            if val != "":
                try:
                    fv  = float(val)
                    val = int(fv) if fv == int(fv) else fv
                except (ValueError, TypeError):
                    pass
            cell = ws.cell(row=er, column=ci, value=val if val != "" else None)
            cell.font=body_font; cell.border=border
            is_id = ci <= len(id_cols)
            cell.alignment = left_al if is_id else centre
            cell.fill      = id_fill if is_id else fill

    # Column widths
    for ci, cn in enumerate(col_order, 1):
        w = 20 if (ci <= len(id_cols) or "magnifications" in cn) else 10
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes = ws.cell(row=3, column=len(id_cols)+1)
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 30

    wb.save(path)
    print(f"  Saved: {path}")


# ---- CLI --------------------------------------------------------------------

def main():
    ap = argparse.ArgumentParser(
        description=(
            "Compute per-sample statistics from an AgNW summary CSV.\n"
            "Pools all magnifications of the same sample into one row.\n"
            "Excludes images with magnification below --min_mag (default 20)."
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    ap.add_argument("input_csv", type=Path,
                    help="Path to all_samples_summary.csv or all_images_summary.csv")
    ap.add_argument("--output", type=Path, default=None,
                    help="Output path (.xlsx or .csv). "
                         "Default: sample_statistics.xlsx next to the input.")
    ap.add_argument("--min_mag", type=int, default=20,
                    help="Minimum plausible magnification (default 20). "
                         "Images below this are excluded.")
    args = ap.parse_args()

    if not args.input_csv.exists():
        print(f"ERROR: File not found: {args.input_csv}")
        sys.exit(1)

    print(f"\nReading : {args.input_csv}")
    print(f"Min mag : {args.min_mag}x  (images below this are excluded)\n")

    fieldnames, rows = build_stats_table(args.input_csv, min_mag=args.min_mag)

    print(f"\n{len(rows)} sample group(s):\n")
    for r in rows:
        parts = [r.get(c,"") for c in ["synthesis_date","sample","modality"] if r.get(c)]
        mags  = r.get("magnifications_included","")
        n     = r.get("n_wires_total","?")
        lm    = r.get("length_um_mean","?"); ls = r.get("length_um_sd","?")
        dm    = r.get("diameter_nm_mean",""); ds = r.get("diameter_nm_sd","")
        dstr  = f"  diam={dm}+/-{ds} nm" if dm else ""
        print(f"  {" | ".join(parts):<40}  mags=[{mags}]  N={n}  len={lm}+/-{ls} um{dstr}")

    out_path = args.output or (args.input_csv.parent / "sample_statistics.xlsx")
    out_path.parent.mkdir(parents=True, exist_ok=True)

    print()
    if out_path.suffix.lower() == ".xlsx":
        save_xlsx(fieldnames, rows, out_path)
    else:
        if out_path.suffix.lower() not in (".csv",):
            out_path = out_path.with_suffix(".csv")
        save_csv(fieldnames, rows, out_path)

    print("\nDone.\n")


if __name__ == "__main__":
    main()
